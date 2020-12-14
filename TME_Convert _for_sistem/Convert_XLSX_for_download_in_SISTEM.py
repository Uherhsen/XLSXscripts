# -*- coding: utf-8 -*-
"""
Created on Wed Nov  4 14:17:16 2020

@author: user

Скрипт для конвертирования файла для отправки в систему.
"""
import openpyxl, os,shutil
from openpyxl import Workbook

def work_sheet(path):
    # Открываем Эксель
    wb = openpyxl.load_workbook(path)#путь к файлу
    sheet = wb.active
    wb.save(path)
    return sheet

# Функция считает все ячейки первого столбца в которых что то написано,до тех пор пока не встретит пустую ячейку "None"
def number_of_articles(sheet):
    """Функция определяет номер первой пустой ячеки в первом столбце файла xlsx и возвращает его"""
    i = 1
    while sheet['A'+str(i)].value != None:
        i+=1
    return i-1

def parse_XLSX(sheet, quantity):
    data=[]
    for i in range(1,quantity+1):
        line_data=[]
        line_data.append(sheet['B'+str(i)].value) # Артикул TME    [0]
        line_data.append(sheet['C'+str(i)].value) # Дескрипшн      [1]
        line_data.append(sheet['D'+str(i)].value) # Вес            [2]
        line_data.append(sheet['E'+str(i)].value) # Картинка       [3]
        line_data.append(sheet['H'+str(i)].value) # PDF            [4]
        line_data.append(sheet['I'+str(i)].value) # Описание       [5]
        line_data.append(sheet['K'+str(i)].value) # Производитель  [6]
        line_data.append(sheet['L'+str(i)].value) # Производитель  [7]
        data.append(line_data)
    return data   
        
def displace(data, kod_ved = ""):
    """Добавляем данные в новый файл в нужном порядке"""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Номер позиции"
    ws['B1'] = "Количество"
    ws['C1'] = "Маркировка детали"
    ws['D1'] = "Цена"
    ws['E1'] = "Сумма"
    ws['F1'] = "Валюта"
    ws['G1'] = "Страна"
    ws['H1'] = "Артикул поставщика"
    ws['I1'] = "Производитель"
    ws['J1'] = "тариф нт"
    ws['K1'] = "Description Eng"
    ws['L1'] = "код тн вэд"
    ws['M1'] = "Описание"
    ws['N1'] = "вес"
    ws['O1'] = "Datasheet"
    ws['P1'] = "Картинка"
    ws['Q1'] = "Нетарифка"
    n = 0
    for i in range(2, (len(data)+2)):
        ws['A'+str(i)] = i-1 # Нумерация
        ws['B'+str(i)] = 1
        ws['C'+str(i)] = data[n][0] # Артикулы
        ws['D'+str(i)] = 1
        ws['E'+str(i)] = 1
        ws['F'+str(i)] = "USD"
        ws['G'+str(i)] = "US"
        ws['H'+str(i)] = data[n][7] # Артикул поставщика
        ws['I'+str(i)] = data[n][6] # Производитель
        ws['K'+str(i)] = data[n][1] # Дескрипшн
        ws['L'+str(i)] = kod_ved # Код ТН ВЭД
        ws['M'+str(i)] = data[n][5] # Описание
        ws['N'+str(i)] = data[n][2] # Вес
        ws['O'+str(i)] = data[n][4] # ДАташит
        ws['P'+str(i)] = data[n][3] # Картинка
        n += 1
    wb.save(path[:-5]+'_convert_for_download.xlsx')
        
path = input("Имя файла xlsx (полностью с расширением): \n")     

kod_tn_ved = input("Код ТН ВЭД который будет проставляться (можно пропустить): \n")

sheet = work_sheet(path)
quantity = number_of_articles(sheet)
my_data = parse_XLSX(sheet, quantity)
displace(my_data, kod_ved=kod_tn_ved) 



  