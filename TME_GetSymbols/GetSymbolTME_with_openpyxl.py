# -*- coding: utf-8 -*-
"""
Created on Tue Jan 14 12:02:49 2020

Получение артикулов по ID с сайта TME через его API (стр )
"""
import openpyxl, os,shutil
from TME_Python_API import product_import_tme


# Функция считает все ячейки первого столбца в которых что то написано,до тех пор пока не встретит пустую ячейку "None"
def number_of_articles(path):
    """Функция определяет номер первой пустой ячеки в первомстолбце файла xlsx"""
    # Открываем Эксель
    wb = openpyxl.load_workbook(path)#путь к файлу
    sheet = wb.active
    #выясняем количество артикулов в файле эксель
    i = 1
    while sheet['A'+str(i)].value != None:
        i+=1
    wb.save(path)
    return i-1

#        
# Функция получает артикулы по ID 
#            
def GetSymbolsList( params, token, app_secret, action1):
    '''Получение артикулов'''
    print('Получение артикулов')
    
    try:
        # Полученные данные
        all_data = product_import_tme(token, app_secret, action1, params)      
        print(all_data['Data']['SymbolList'])
        SymbolList = all_data['Data']['SymbolList'] # получаем список артикулов по ID
        return SymbolList
    except IndexError:
        print("Ошибка структуры ответа")
    
def fill_in_XLSX(path, params,token, app_secret, action1):
    """Заполнение файла symbolsdata.xlsx артикулами"""
    # Открываем Эксель
    wb = openpyxl.load_workbook(path) #путь к файлу
    sheet = wb.active
    
    rng0 = number_of_articles(path) + 1
    SymbolList = GetSymbolsList( params,token, app_secret, action1)
    
    for i in SymbolList:
        sheet['A'+str(rng0)] = i
        rng0 = rng0 + 1
        print(rng0)
    wb.save(path)
    print('ГОТОВО\n')
    return rng0-1
    
def delAll():
    """Функция отчистки файла symbolsdata.xlsx от данных"""
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    wb.save( "symbolsdata.xlsx")
    
def new_name_copy(xlsxpath,pcs,noid):
    """Создание файла с новым именем, в название файла добавляется количество позиций и id ветки. Аргументы: путь к файлу, кол-во позий, номер ветктки - id"""
    xlsxpath_copy = xlsxpath[:-5] + '_'+ str(pcs)  + 'pcs_id' + str(noid) + '.xlsx'
    if os.path.exists(xlsxpath_copy):
        os.remove(xlsxpath_copy)
        shutil.copy2(xlsxpath, xlsxpath_copy)
    else:
        shutil.copy2(xlsxpath, xlsxpath_copy)
        
if __name__ == '__main__': 

    xlsxpath = "symbolsdata.xlsx"
    
    print("Требуется ввести ID ветки продуктов с ТМЕ")
    produkt_ID = input("Для выхода введи n\n# ID: ")
    
    params={'CategoryId' : produkt_ID, 'Country' : 'RU','Language' : 'RU',} #100544(резисторы-сделаны), 112141(диоды),112807(мостовые выпрямители), 19(тиристоры), tranzistory_112825,
    token = 'ac434c181917ed4e51c49a2027bfd040e9f2da0054be7'
    app_secret = '0b748f6e5d340d693703'
    action1 = 'Products/GetSymbols' # request method, метод пинг Utils/Ping
    while produkt_ID != "n":
        params={'CategoryId' : produkt_ID, 'Country' : 'RU','Language' : 'RU',}
        pcs = fill_in_XLSX(xlsxpath, params, token, app_secret, action1)
        new_name_copy(xlsxpath,pcs,params['CategoryId']) 
        delAll()
        print("Требуется ввести ID ветки продуктов с ТМЕ")
        produkt_ID = input("Для выхода введи n\n# ID: ")
    else:
        print("отмена")